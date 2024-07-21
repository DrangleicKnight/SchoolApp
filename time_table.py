import openpyxl
from kivy.metrics import dp
from kivymd.uix.datatables import MDDataTable


def time_table(monday1, monday2, monday3, monday4, monday5, monday6,
               tuesday1, tuesday2, tuesday3, tuesday4, tuesday5, tuesday6,
               wednesday1, wednesday2, wednesday3, wednesday4, wednesday5, wednesday6,
               thursday1, thursday2, thursday3, thursday4, thursday5, thursday6,
               friday1, friday2, friday3, friday4, friday5, friday6,
               saturday1, saturday2):
    workbook = openpyxl.load_workbook("Time Table.xlsx")
    worksheet = workbook.active

    cols = worksheet.iter_cols(min_row=4, max_row=9,
                               min_col=2, max_col=worksheet.max_column)

    for entries in cols:
        for i in range(len(entries)):
            if entries[i].value != "Break":
                entries[i].value = " "


    worksheet["B4"] = monday1.text
    worksheet["C4"] = monday2.text
    worksheet["E4"] = monday3.text
    worksheet["F4"] = monday4.text
    worksheet["H4"] = monday5.text
    worksheet["I4"] = monday6.text

    worksheet["B5"] = tuesday1.text
    worksheet["C5"] = tuesday2.text
    worksheet["E5"] = tuesday3.text
    worksheet["F5"] = tuesday4.text
    worksheet["H5"] = tuesday5.text
    worksheet["I5"] = tuesday6.text

    worksheet["B6"] = wednesday1.text
    worksheet["C6"] = wednesday2.text
    worksheet["E6"] = wednesday3.text
    worksheet["F6"] = wednesday4.text
    worksheet["H6"] = wednesday5.text
    worksheet["I6"] = wednesday6.text

    worksheet["B7"] = thursday1.text
    worksheet["C7"] = thursday2.text
    worksheet["E7"] = thursday3.text
    worksheet["F7"] = thursday4.text
    worksheet["H7"] = thursday5.text
    worksheet["I7"] = thursday6.text

    worksheet["B8"] = friday1.text
    worksheet["C8"] = friday2.text
    worksheet["E8"] = friday3.text
    worksheet["F8"] = friday4.text
    worksheet["H8"] = friday5.text
    worksheet["I8"] = friday6.text

    worksheet["B9"] = saturday1.text
    worksheet["C9"] = saturday2.text

    workbook.save("Time Table.xlsx")
    workbook.close()


def appoint_teachers(python_teacher, devops_teacher, os_teacher, android_teacher, project_teacher):
    workbook = openpyxl.load_workbook("Time Table.xlsx")
    worksheet = workbook.active

    worksheet["F15"] = python_teacher.text
    worksheet["F16"] = devops_teacher.text
    worksheet["F17"] = os_teacher.text
    worksheet["F18"] = android_teacher.text
    worksheet["F19"] = project_teacher.text

    workbook.save("Time Table.xlsx")
    workbook.close()


def datatable(self):
    workbook = openpyxl.load_workbook("Time Table.xlsx")
    worksheet = workbook.active

    data = []

    for rows in worksheet.iter_rows(min_col=2, max_col=9, min_row=4, max_row=9):
        for i in range(8):
            data.append(rows[i].value)

    workbook.close()

    table = MDDataTable(background_color = "4E3636",
                        background_color_header = "#B8621B",
                        rows_num = 6,
                        column_data=[("[size=24]   Day[/size]", dp(25)),
                                     ("9:00 - 10:00", dp(25)),
                                     ("10:00 - 11:00", dp(25)),
                                     ("11:15 - 12:15", dp(25)),
                                     ("11:15 - 12:15", dp(25)),
                                     ("12:15 - 1:15", dp(25)),
                                     ("1:15 - 2:15", dp(25)),
                                     ("2:15 - 3:15", dp(25)),
                                     ("3:15 - 4:15", dp(25))],
                        row_data = [("[size=24]  Monday[/size]", data[0], data[1], "[color=#FE0000]Break[/color]", data[3], data[4], "[color=#FE0000]Break[/color]", data[6], data[7]),
                                    ("[size=24]  Tuesday[/size]", data[8], data[9], "[color=#FE0000]Break[/color]", data[11], data[12], "[color=#FE0000]Break[/color]", data[14], data[15]),
                                    ("[size=24]  Wednesday[/size]", data[16], data[17], "[color=#FE0000]Break[/color]", data[19], data[20], "[color=#FE0000]Break[/color]", data[22], data[23]),
                                    ("[size=24]  Thursday[/size]", data[24], data[25], "[color=#FE0000]Break[/color]", data[27], data[28], "[color=#FE0000]Break[/color]", data[30], data[31]),
                                    ("[size=24]  Friday[/size]", data[32], data[33], "[color=#FE0000]Break[/color]", data[35], data[36], "[color=#FE0000]Break[/color]", data[38], data[39]),
                                    ("[size=24]  Saturday[/size]", data[40], data[41], " ", " ", " ", " ", " ", " ")])

    screen = self.root.get_screen("classroom").ids.time_table
    screen.add_widget(table)