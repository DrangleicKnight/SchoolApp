import os
import numpy as np
import face_recognition
import cv2
import openpyxl
from openpyxl.styles import PatternFill

def face_encoder(path, faces):
    face = face_recognition.load_image_file(path)
    encoding = face_recognition.face_encodings(face)[0]
    faces.append(encoding)

def studentDetect(self, date):
    video_capture = cv2.VideoCapture(0, cv2.CAP_DSHOW)
    video_capture.set(cv2.CAP_PROP_FPS, 30)
    video_capture.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc(*'MJPG'))

    dirs = []
    names = []

    for imagess in os.listdir("E:\\MiniProject - Students Management\\images\\students\\users"):
        dirs.append(f"images/students/users/{imagess}")
        names.append(imagess.split(".")[0])

    known_face_encodings = []
    known_face_names = [name for name in names]

    for i in range(len(dirs)):
        face_encoder(dirs[i], known_face_encodings)

    face_locations = []
    face_encodings = []
    face_names = []
    full_info = []
    process_this_frame = True

    while video_capture:
        ret, frame = video_capture.read()

        if process_this_frame:
            small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
            rgb_small_frame = small_frame[:, :, ::-1]

            face_locations = face_recognition.face_locations(rgb_small_frame)
            face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

            face_names = []
            for face_encoding in face_encodings:
                matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
                name = "Unknown"

                face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
                best_match_index = np.argmin(face_distances)
                if matches[best_match_index]:
                    name = known_face_names[best_match_index]

                face_names.append(name)

                workbook = openpyxl.load_workbook("Attendance.xlsx")
                sheet = workbook.active

                rows = sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=3)
                cols = sheet.iter_cols(min_col=4, max_col=sheet.max_column, min_row=2, max_row=sheet.max_row)

                student_list = []

                for roll, namesake, classed in rows:
                    student_list.append(namesake.value)

                for attendance in cols:
                    for i in range(len(attendance)):
                        if student_list.__contains__(name) and attendance[i].value == date.text:
                            attendance[student_list.index(name) + 1].value = "Present"
                            attendance[student_list.index(name) + 1].fill = PatternFill(start_color="7AA874",
                                                                                        end_color="7AA874",
                                                                                        fill_type="solid")
                            break
                        break
                workbook.save("Attendance.xlsx")
                workbook.close()

        process_this_frame = not process_this_frame

        for (top, right, bottom, left), name in zip(face_locations, face_names):
            top *= 4
            right *= 4
            bottom *= 4
            left *= 4

            cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)

            cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
            font = cv2.FONT_HERSHEY_DUPLEX

            cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)

        cv2.imshow('Video', frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    video_capture.release()
    cv2.destroyAllWindows()