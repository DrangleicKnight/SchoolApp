import openpyxl
from openpyxl.styles import PatternFill


def load_data(self):
    workbook = openpyxl.load_workbook("Attendance.xlsx")
    sheet = workbook.active

    rows = sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=3)

    student_roll = []
    student_name = []
    student_class = []

    student_roll_xl = []
    student_name_xl = []

    for student in self.studentsCollection.find():
        roll_no = int(student.get("Roll Number"))
        name = student.get("Name")
        classed = student.get("Class")
        student_roll.append(roll_no)
        student_name.append(name)
        student_class.append(classed)

    for roll, name, classed in rows:
        student_roll_xl.append(roll.value)
        student_name_xl.append(name.value)

    for i in range(len(student_roll)):
        if student_roll[i] not in student_roll_xl:
            sheet.append([student_roll[i], student_name[i], student_class[i]])

    workbook.save("Attendance.xlsx")
    workbook.close()
    set_xl_data(student_roll)

def set_xl_data(student_roll):
    workbook = openpyxl.load_workbook("Attendance.xlsx")
    sheet = workbook.active

    cols = sheet.iter_cols(min_col=4, max_col=sheet.max_column, min_row=3, max_row=sheet.max_row)
    for attendance in cols:
        for i in range(len(student_roll)):
            if not attendance[i].value:
                attendance[i].value = "Absent"
                attendance[i].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    workbook.save("Attendance.xlsx")
    workbook.close()
